import os
from openpyxl import load_workbook
import arcgis

import os
import mimetypes
import sys
import django
import requests
import logging
import json
import traceback
import pyodbc
import datetime
import xlrd
from arcgis.gis import GIS
from arcgis.features import FeatureLayerCollection
import pprint
import urllib.request
import urllib.parse
from operator import itemgetter

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
os.environ['DJANGO_SETTINGS_MODULE'] = 'rtaa_gis.settings'
django.setup()

##############################################
if not os.path.exists(os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")):
    os.mkdir(os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs"))

log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs/Excel_Domain_bridge.txt")
file = open(log_path, 'w')
file.close()


def loggit(text):
    file = open(log_path, 'a')
    file.write("{}\n".format(text))
    pprint.pprint("{}\n".format(text))
    file.close()


if __name__ == "__main__":
    try:
        wb = load_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'RTAA_Domain_Sheets.xlsx'),
                           data_only=True)
        sheet_names = wb.get_sheet_names()

        # send post request to update the domains in AGOL
        token_url = r"https://www.arcgis.com/sharing/rest/generateToken"
        params = {
            'f': 'pjson',
            'username': 'data_owner',
            'password': 'GIS@RTAA123!',
            'referer': 'http://www.arcgis.com'
        }
        data = urllib.parse.urlencode(params)
        data = data.encode('ascii')

        req = urllib.request.Request(token_url, data)
        response = urllib.request.urlopen(req)
        data = response.read().decode("utf-8")
        # Convert string to dictionary
        json_acceptable_string = data.replace("'", "\"")
        d = json.loads(json_acceptable_string)
        token = d['token']

        # Query the tables and update the data in AGOL
        gis = GIS("https://www.arcgis.com", "data_owner", "GIS@RTAA123!")
        layer = gis.content.get('290de3849d6c412f950471bc45df182a')

        feature_layer = layer.layers[0]
        # Update the domains for the feature service

        existing_fields = feature_layer.properties["fields"]
        new_fields = [x for x in existing_fields]
        for obj in new_fields:
            # for each field, if it matches a sheet name, build the domain from the excel sheet
            field_name = obj["name"]
            if field_name.lower() in [x.lower().replace(" ", "_") for x in sheet_names]:
                empty_list = []
                domain_list = []
                ws = wb.get_sheet_by_name(field_name.lower().replace("_", " "))
                for row in ws.iter_rows(min_row=2):
                    code = row[0].value

                    val = {
                        "code": "{}".format(code),
                        "name": "{}".format(code)
                    }
                    empty_list.append(val)

                # sort the keys, then build the final sorted domain list
                new_domains = sorted(empty_list, key=itemgetter('name'))
                domain_list.extend(new_domains)
                if not obj["domain"]:
                    # the domain is null so create one
                    domain_type = "codedValue"
                    domain_name = "Code{}".format(field_name.replace("_", "").strip())
                    obj["domain"] = {
                        "type": domain_type,
                        "name": domain_name,
                        "codedValues": domain_list
                    }

                else:
                    # Otherwise just update the coded values
                    domain_type = obj["domain"]["type"]
                    domain_name = obj["domain"]["name"]
                    obj["domain"] = {
                        "type": domain_type,
                        "name": domain_name,
                        "codedValues": domain_list
                    }

        pprint.pprint(new_fields)

        post_data = {
            "token": token,
            "f": "pjson",
            "updateDefinition": {
                "fields": new_fields
            }
        }

        # the constant types in python must be converted to null and true or false
        d = urllib.parse.urlencode(post_data).replace('None', 'null').replace('False', 'false').replace('True', 'true')
        post_data = d.encode('ascii')
        # the admin service url must be used
        service_url = r"https://services6.arcgis.com/GC5xdlDk2dEYWofH/arcgis/rest/admin/services/Space/FeatureServer/0/updateDefinition"
        req = urllib.request.Request(service_url, data=post_data, method="POST")
        response = urllib.request.urlopen(req)
        data = response.read().decode("utf-8")
        json_acceptable_string = data.replace("'", "\"")
        d = json.loads(json_acceptable_string)
        pprint.pprint(d)

    except Exception as e:
        traceback.print_exc(file=sys.stdout)
        loggit(e)

